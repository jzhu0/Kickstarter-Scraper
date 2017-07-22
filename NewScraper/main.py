'''
Created on Jun 21, 2017
Scrapes through the Kickstarter Longitudinal data set, grabbing the contents of the project
homepage, updates, and creator bio.

Assumes that column A of the data set is "Project Url", and column B contains the Project ID (as an int).
Output data is stored with Project IDs in alphabetically sorted order (can re-order within Excel),
and lots of random whitespace within the homepage/updates/bio columns.

@author: Jason Zhu
'''
DATASET_NAME = "COPY - Kickstarter Longitudinal - 6.13.2017.xlsx"
DATASET_SHEETNAME = "Kickstarter V.6"
NEW_DATASET_FILENAME = "Founder bios & Campaign text - 6.22.2017.xlsx"

import time
from bs4 import BeautifulSoup
import requests
import openpyxl
from selenium import webdriver

# Extraction functions
def get_homepage(data):
    homepage = data.find("div", {"class": "col col-8 description-container"}).text
    return homepage

def get_updates(data):
    updates = data.find("div", {"class" : "timeline"}).text
    return updates

def get_bio(data):
    bio = data.find_all("div", {"class" : "col-full col-sm-15-20 col-md-11-16"})
    if (len(bio) == 0):   # case where profile is deleted
        return "[DELETED PROFILE]"
    return bio[0].text

def check_deleted(data):
    deleted = data.find_all("div", {"id" : "purged_project"})
    return len(deleted) != 0

print "Loading data set " + DATASET_NAME + "..."
KickLong = openpyxl.load_workbook(DATASET_NAME)
links = {}
if (DATASET_SHEETNAME in KickLong.sheetnames):
    KickstarterV6 = KickLong[DATASET_SHEETNAME]
else:
    raise Exception("Data set did not contain correct data sheet")
print "    Data set successfully opened"
for row in KickstarterV6.iter_rows():
    if (row[0].value == "Project Url"):   # skip the first row
        continue
    if not (row[1].value in links):
        links[row[1].value] = str(row[0].value)
        print "Loaded pID #" + str(row[1].value)
#    if (row[1].value > 20):   # test loading a few rows
#        break
print "    Data set successfully loaded"

#test with smaller links dict:
#print "Using hard-coded data set."
#links = {1 : "https://www.kickstarter.com/projects/1000809202/hydroback-hydration-systems",
#6 : "https://www.kickstarter.com/projects/1006232434/worlds-first-battery-powered-bike-pump",
#9 : "https://www.kickstarter.com/projects/1012933739/monster-towel-0",
#14 : "https://www.kickstarter.com/projects/1019342199/meet-flipback-the-shoulder-rest-cell-phone-case"}

#print "Using hard-coded annoying links set."
#links = {269 : "https://www.kickstarter.com/projects/1550173720/pinthusiasts-the-ultimate-app-in-pin-trading-and-s",
#326 : "https://www.kickstarter.com/projects/1674004085/isociometrics-social-media-analytics-for-your-bran"}

driver = webdriver.Chrome()

# Set up new output Excel sheet
new_data = openpyxl.Workbook()
new_data1 = new_data.active
new_data1.title = "Sheet 1"
new_data1['A1'] = "Project ID"
new_data1['B1'] = "Homepage"
new_data1.column_dimensions['B'].width = 50
new_data1['C1'] = "Updates"
new_data1.column_dimensions['C'].width = 50
new_data1['D1'] = "Creator Bio"
new_data1.column_dimensions['D'].width = 50

row = 2
for pID, pURL in links.iteritems():
    print "Processing pID #" + str(pID)
    req = requests.get(pURL)
    data = BeautifulSoup(req.content, "html.parser")
    new_data1['A' + str(row)] = pID
    if (check_deleted(data)):
        print "  pID #" + str(pID) + " is no longer available."
        new_data1['B' + str(row)] = "[DELETED PROJECT]"
        new_data1['C' + str(row)] = "[DELETED PROJECT]"
        new_data1['D' + str(row)] = "[DELETED PROJECT]"
    else:
        # Get homepage 
        time.sleep(0.5)
        homepage = get_homepage(data)
        new_data1['B' + str(row)] = homepage
    
        # Get updates
        req = requests.get(pURL + "/updates")
        data = BeautifulSoup(req.content, "html.parser")
        time.sleep(0.5)
        updates = get_updates(data)
        new_data1['C' + str(row)] = updates
    
        # Get creator bio
        driver.get(pURL + "/creator_bio")
        profileLink = driver.find_element_by_class_name("green-dark")
        profileUrl = profileLink.get_attribute("href")
        req = requests.get(profileUrl + "/about")
        data = BeautifulSoup(req.content, "html.parser")
        time.sleep(0.5)
        bio = get_bio(data)
        new_data1['D' + str(row)] = bio
    
    row += 1
    
new_data.save(filename = NEW_DATASET_FILENAME)
print "    All data successfully parsed."
